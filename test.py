import pandas as pd
import os
import json
import psycopg2
import openpyxl
import pprint

# Ouvre le fichier Excel
wb = openpyxl.load_workbook('maquette4A.xlsx')
ws = wb.active

# Lire le contenu de la cellule A8
code_ligne_8 = ws['A8'].value

# Initialisation
data = {}
current_sem = None
current_ue = None

# Fonction générique de traitement
def traiter_lignes(ue_types):
    global data, current_sem, current_ue
    data = {}
    current_sem = None
    current_ue = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        code = row[0]
        type_element = row[4]
        nom = row[2]
        coeff = row[5] if row[5] else 0

        if type_element == 'SEM':
            current_sem = {'code': code, 'nom': nom, 'ues': []}
            data[nom] = current_sem

        elif type_element in ue_types:
            current_ue = {'code': code, 'nom': nom, 'ecues': []}
            current_sem['ues'].append(current_ue)

        elif type_element == 'ECUE':
            ecue_data = {
                'code': code,
                'nom': nom,
                'coeff': float(coeff)
            }
            current_ue['ecues'].append(ecue_data)

        elif type_element == 'CALC':
            ue_coeff = sum(ecue['coeff'] for ecue in current_ue['ecues'])
            calc_data = {
                'code': code,
                'nom': nom,
                'coeff': -ue_coeff
            }
            current_ue['ecues'].append(calc_data)

# Choix du traitement selon le code trouvé en A8
if code_ligne_8 == 'JGI5SAA':
    traiter_lignes(ue_types={'UE'})  # UE uniquement
elif code_ligne_8 == 'JGI7SAA':
    traiter_lignes(ue_types={'UE', 'STAG'})  # UE et STAG
else:
    print(f"Code inconnu dans A8 : {code_ligne_8}")

# Connexion à la base de données PostgreSQL
conn = psycopg2.connect(
    dbname='gii_polytech',
    user='postgres',
    password='postgres',
    host='localhost',
    port='5432'
)
cur = conn.cursor()

# Insertion sans fournir l'ID (géré par SERIAL/auto-incrément)
for semestre_code, semestre in data.items():
    cur.execute("INSERT INTO semestre (code_semestre, nom_semestre) VALUES (%s, %s) RETURNING id_semestre",
                (semestre["code"], semestre["nom"]))
    semestre_id = cur.fetchone()[0]

    for ue in semestre['ues']:
        cur.execute("INSERT INTO ue (code_ue, nom_ue, id_semestre) VALUES (%s, %s, %s) RETURNING id_ue",
                    (ue["code"], ue["nom"], semestre_id))
        ue_id = cur.fetchone()[0]

        for ecue in ue["ecues"]:
            cur.execute("INSERT INTO ecue (code_ecue, nom_ecue, coefficient, id_ue) VALUES (%s, %s, %s, %s)",
                        (ecue["code"], ecue["nom"], ecue["coeff"], ue_id))

conn.commit()
cur.close()
conn.close()

print("Insertion terminée avec succès.")
